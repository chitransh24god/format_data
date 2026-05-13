"""
DataForge — Universal Converter Engine
Detects file type, cleans data, and outputs structured Excel workbooks.
"""
from __future__ import annotations

import io
import json
import re
import traceback
import warnings
from pathlib import Path
from typing import Any

import chardet
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ── Type detection helpers ────────────────────────────────────────────────────

DELIMITERS = [",", "\t", ";", "|", ":", "  "]
EXCEL_EXTS = {".xlsx", ".xls", ".xlsm", ".ods"}
TEXT_EXTS = {".txt", ".log", ".text"}
SQL_KEYWORDS = re.compile(
    r"\b(SELECT|INSERT|UPDATE|DELETE|CREATE|DROP|ALTER|FROM|WHERE|VALUES)\b",
    re.IGNORECASE,
)


def detect_encoding(raw: bytes) -> str:
    result = chardet.detect(raw[:50_000])
    return result.get("encoding") or "utf-8"


def sniff_delimiter(text: str) -> str:
    first_lines = "\n".join(text.splitlines()[:10])
    counts = {d: first_lines.count(d) for d in DELIMITERS}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def is_json(text: str) -> bool:
    t = text.strip()
    return (t.startswith("{") or t.startswith("[")) and (t.endswith("}") or t.endswith("]"))


def is_xml(text: str) -> bool:
    return text.strip().startswith("<") and "</" in text


def is_yaml(text: str) -> bool:
    lines = [l for l in text.splitlines() if l.strip() and not l.strip().startswith("#")]
    if not lines:
        return False
    return any(":" in l and not l.startswith("<") for l in lines[:5])


def is_html_table(text: str) -> bool:
    return bool(re.search(r"<table", text, re.IGNORECASE))


def is_sql_dump(text: str) -> bool:
    return bool(SQL_KEYWORDS.search(text[:2000]))


def is_key_value(text: str) -> bool:
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        return False
    kv = sum(1 for l in lines[:20] if "=" in l or ": " in l)
    return kv / max(len(lines[:20]), 1) > 0.6


# ── Readers ───────────────────────────────────────────────────────────────────

def read_bytes(uploaded) -> tuple[bytes, str]:
    """Return raw bytes from an uploaded file or path."""
    if hasattr(uploaded, "read"):
        raw = uploaded.read()
        name = getattr(uploaded, "name", "file")
        uploaded.seek(0)
    else:
        raw = Path(uploaded).read_bytes()
        name = Path(uploaded).name
    return raw, name


def parse_json(text: str, name: str) -> dict[str, pd.DataFrame]:
    data = json.loads(text)
    sheets: dict[str, pd.DataFrame] = {}

    def _flatten(obj, prefix="") -> dict:
        out = {}
        if isinstance(obj, dict):
            for k, v in obj.items():
                key = f"{prefix}.{k}" if prefix else k
                if isinstance(v, (dict, list)):
                    out.update(_flatten(v, key))
                else:
                    out[key] = v
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                out.update(_flatten(item, f"{prefix}[{i}]"))
        else:
            out[prefix] = obj
        return out

    if isinstance(data, list):
        if all(isinstance(r, dict) for r in data):
            sheets["Data"] = pd.json_normalize(data)
        else:
            sheets["Data"] = pd.DataFrame({"value": data})
    elif isinstance(data, dict):
        # Check if any values are lists of dicts (nested tables)
        for k, v in data.items():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                sheets[str(k)[:31]] = pd.json_normalize(v)
        if not sheets:
            sheets["Data"] = pd.DataFrame([_flatten(data)])
    return sheets


def parse_jsonl(text: str) -> dict[str, pd.DataFrame]:
    records = [json.loads(l) for l in text.splitlines() if l.strip()]
    return {"Data": pd.json_normalize(records)}


def parse_xml(raw: bytes, name: str) -> dict[str, pd.DataFrame]:
    try:
        frames = pd.read_xml(io.BytesIO(raw))
        return {"Data": frames}
    except Exception:
        import xml.etree.ElementTree as ET
        root = ET.fromstring(raw)
        records = []
        for child in root:
            rec = child.attrib.copy()
            for sub in child:
                rec[sub.tag] = sub.text
            records.append(rec)
        return {"Data": pd.DataFrame(records) if records else pd.DataFrame()}


def parse_yaml(text: str) -> dict[str, pd.DataFrame]:
    import yaml
    data = yaml.safe_load(text)
    if isinstance(data, list):
        return {"Data": pd.json_normalize(data)}
    elif isinstance(data, dict):
        sheets = {}
        for k, v in data.items():
            if isinstance(v, list):
                sheets[str(k)[:31]] = pd.json_normalize(v)
        return sheets or {"Data": pd.DataFrame([data])}
    return {"Data": pd.DataFrame()}


def parse_html(raw: bytes) -> dict[str, pd.DataFrame]:
    tables = pd.read_html(io.BytesIO(raw))
    return {f"Table_{i+1}": df for i, df in enumerate(tables)} if tables else {}


def parse_sql_dump(text: str) -> dict[str, pd.DataFrame]:
    """Extract INSERT INTO statements from SQL dumps."""
    pattern = re.compile(
        r"INSERT\s+INTO\s+[`'\"]?(\w+)[`'\"]?\s*\(([^)]+)\)\s*VALUES\s*(.+?)(?:;|$)",
        re.IGNORECASE | re.DOTALL,
    )
    sheets: dict[str, list] = {}
    for match in pattern.finditer(text):
        table = match.group(1)
        cols = [c.strip().strip("`'\"") for c in match.group(2).split(",")]
        vals_str = match.group(3)
        row_matches = re.findall(r"\(([^)]+)\)", vals_str)
        for row_str in row_matches:
            vals = [v.strip().strip("'\"") for v in row_str.split(",")]
            if table not in sheets:
                sheets[table] = []
            sheets[table].append(dict(zip(cols, vals)))
    return {k[:31]: pd.DataFrame(v) for k, v in sheets.items()} if sheets else {"SQL": pd.DataFrame()}


def parse_parquet(raw: bytes) -> dict[str, pd.DataFrame]:
    return {"Data": pd.read_parquet(io.BytesIO(raw))}


def parse_feather(raw: bytes) -> dict[str, pd.DataFrame]:
    return {"Data": pd.read_feather(io.BytesIO(raw))}


def parse_hdf(path_or_buf, name: str) -> dict[str, pd.DataFrame]:
    with pd.HDFStore(path_or_buf, "r") as store:
        keys = store.keys()
    sheets = {}
    for k in keys:
        label = k.lstrip("/")[:31]
        sheets[label] = pd.read_hdf(path_or_buf, key=k)
    return sheets


def parse_csv(raw: bytes, delimiter: str | None = None) -> dict[str, pd.DataFrame]:
    enc = detect_encoding(raw)
    text = raw.decode(enc, errors="replace")
    sep = delimiter or sniff_delimiter(text)
    df = pd.read_csv(io.StringIO(text), sep=sep, engine="python", on_bad_lines="skip")
    return {"Data": df}


def parse_excel(raw: bytes, ext: str) -> dict[str, pd.DataFrame]:
    engine = "openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd" if ext == ".xls" else "odf"
    xf = pd.ExcelFile(io.BytesIO(raw), engine=engine)
    return {sheet[:31]: xf.parse(sheet) for sheet in xf.sheet_names}


def parse_pdf(raw: bytes) -> dict[str, pd.DataFrame]:
    try:
        import pdfplumber
        all_rows: list[dict] = []
        table_sheets: dict[str, pd.DataFrame] = {}
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for i, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                for j, tbl in enumerate(tables):
                    if tbl and len(tbl) > 1:
                        df = pd.DataFrame(tbl[1:], columns=tbl[0])
                        key = f"P{i+1}_T{j+1}"
                        table_sheets[key] = df
                if not tables:
                    text = page.extract_text() or ""
                    for line in text.splitlines():
                        if line.strip():
                            all_rows.append({"page": i + 1, "text": line.strip()})
        if table_sheets:
            return table_sheets
        return {"Text": pd.DataFrame(all_rows)}
    except ImportError:
        return {"Error": pd.DataFrame([{"message": "Install pdfplumber: pip install pdfplumber"}])}


def parse_key_value(text: str) -> dict[str, pd.DataFrame]:
    records = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, _, v = line.partition("=")
        elif ": " in line:
            k, _, v = line.partition(": ")
        else:
            continue
        records[k.strip()] = v.strip()
    df = pd.DataFrame(list(records.items()), columns=["Key", "Value"])
    return {"Properties": df}


def parse_text_log(text: str) -> dict[str, pd.DataFrame]:
    """Try structured log parsing, fall back to line-by-line."""
    # Common log pattern: timestamp level message
    log_re = re.compile(
        r"^(?P<timestamp>\d{4}[-/]\d{2}[-/]\d{2}[T ]\d{2}:\d{2}:\d{2}[^\s]*)\s+"
        r"(?P<level>DEBUG|INFO|WARNING|WARN|ERROR|CRITICAL|FATAL)?\s*"
        r"(?P<message>.+)$",
        re.IGNORECASE,
    )
    records = []
    unstructured = []
    for line in text.splitlines():
        m = log_re.match(line.strip())
        if m:
            records.append(m.groupdict())
        else:
            unstructured.append({"line": line})
    if records:
        return {"Logs": pd.DataFrame(records)}
    return {"Lines": pd.DataFrame(unstructured)}


def parse_avro(raw: bytes) -> dict[str, pd.DataFrame]:
    try:
        import fastavro
        reader = fastavro.reader(io.BytesIO(raw))
        return {"Data": pd.DataFrame(list(reader))}
    except ImportError:
        return {"Error": pd.DataFrame([{"message": "Install fastavro: pip install fastavro"}])}


# ── Main dispatcher ───────────────────────────────────────────────────────────

def convert(uploaded, extra_text: str = "") -> dict[str, Any]:
    """
    Convert any uploaded file or raw text to a dict of DataFrames.

    Returns:
        {
            "sheets": dict[str, pd.DataFrame],
            "source_type": str,
            "file_name": str,
            "warnings": list[str],
            "errors": list[str],
        }
    """
    result: dict[str, Any] = {"sheets": {}, "source_type": "unknown", "file_name": "", "warnings": [], "errors": []}

    try:
        # ── Raw text paste ────────────────────────────────────────────────────
        if uploaded is None and extra_text.strip():
            text = extra_text.strip()
            result["file_name"] = "pasted_data"
            if is_json(text):
                result["source_type"] = "JSON"
                result["sheets"] = parse_json(text, "pasted")
            elif is_xml(text):
                result["source_type"] = "XML"
                result["sheets"] = parse_xml(text.encode(), "pasted")
            elif is_yaml(text):
                result["source_type"] = "YAML"
                result["sheets"] = parse_yaml(text)
            elif is_html_table(text):
                result["source_type"] = "HTML"
                result["sheets"] = parse_html(text.encode())
            elif is_sql_dump(text):
                result["source_type"] = "SQL"
                result["sheets"] = parse_sql_dump(text)
            elif is_key_value(text):
                result["source_type"] = "Key-Value"
                result["sheets"] = parse_key_value(text)
            else:
                # Try CSV
                try:
                    delim = sniff_delimiter(text)
                    df = pd.read_csv(io.StringIO(text), sep=delim, engine="python", on_bad_lines="skip")
                    if df.shape[1] > 1:
                        result["source_type"] = f"Delimited ({repr(delim)})"
                        result["sheets"] = {"Data": df}
                    else:
                        raise ValueError("single column")
                except Exception:
                    result["source_type"] = "Plain Text"
                    result["sheets"] = parse_text_log(text)
            return _post_process(result)

        # ── File upload ───────────────────────────────────────────────────────
        raw, name = read_bytes(uploaded)
        result["file_name"] = name
        ext = Path(name).suffix.lower()
        enc = detect_encoding(raw)

        # Excel family
        if ext in EXCEL_EXTS:
            result["source_type"] = f"Excel ({ext})"
            result["sheets"] = parse_excel(raw, ext)

        elif ext == ".csv":
            result["source_type"] = "CSV"
            result["sheets"] = parse_csv(raw)

        elif ext == ".tsv":
            result["source_type"] = "TSV"
            result["sheets"] = parse_csv(raw, delimiter="\t")

        elif ext == ".json":
            text = raw.decode(enc, errors="replace")
            result["source_type"] = "JSON"
            result["sheets"] = parse_json(text, name)

        elif ext == ".jsonl":
            text = raw.decode(enc, errors="replace")
            result["source_type"] = "JSONL"
            result["sheets"] = parse_jsonl(text)

        elif ext == ".xml":
            result["source_type"] = "XML"
            result["sheets"] = parse_xml(raw, name)

        elif ext in (".yaml", ".yml"):
            text = raw.decode(enc, errors="replace")
            result["source_type"] = "YAML"
            result["sheets"] = parse_yaml(text)

        elif ext == ".html" or ext == ".htm":
            result["source_type"] = "HTML"
            result["sheets"] = parse_html(raw)

        elif ext == ".pdf":
            result["source_type"] = "PDF"
            result["sheets"] = parse_pdf(raw)

        elif ext == ".parquet":
            result["source_type"] = "Parquet"
            result["sheets"] = parse_parquet(raw)

        elif ext == ".feather":
            result["source_type"] = "Feather"
            result["sheets"] = parse_feather(raw)

        elif ext in (".h5", ".hdf5", ".hdf"):
            import tempfile, os
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
            try:
                result["source_type"] = "HDF5"
                result["sheets"] = parse_hdf(tmp_path, name)
            finally:
                os.unlink(tmp_path)

        elif ext in (".pkl", ".pickle"):
            obj = pd.read_pickle(io.BytesIO(raw))
            if isinstance(obj, pd.DataFrame):
                result["sheets"] = {"Data": obj}
            elif isinstance(obj, dict):
                result["sheets"] = {str(k)[:31]: v for k, v in obj.items() if isinstance(v, pd.DataFrame)}
            result["source_type"] = "Pickle"

        elif ext == ".avro":
            result["source_type"] = "Avro"
            result["sheets"] = parse_avro(raw)

        elif ext == ".sql":
            text = raw.decode(enc, errors="replace")
            result["source_type"] = "SQL Dump"
            result["sheets"] = parse_sql_dump(text)

        elif ext in TEXT_EXTS or ext == "":
            text = raw.decode(enc, errors="replace")
            # Re-sniff content
            if is_json(text):
                result["source_type"] = "JSON (txt)"
                result["sheets"] = parse_json(text, name)
            elif is_xml(text):
                result["source_type"] = "XML (txt)"
                result["sheets"] = parse_xml(raw, name)
            elif is_yaml(text):
                result["source_type"] = "YAML (txt)"
                result["sheets"] = parse_yaml(text)
            elif is_sql_dump(text):
                result["source_type"] = "SQL (txt)"
                result["sheets"] = parse_sql_dump(text)
            elif is_key_value(text):
                result["source_type"] = "Key-Value"
                result["sheets"] = parse_key_value(text)
            else:
                delim = sniff_delimiter(text)
                try:
                    df = pd.read_csv(io.StringIO(text), sep=delim, engine="python", on_bad_lines="skip")
                    if df.shape[1] > 1:
                        result["source_type"] = f"Delimited ({repr(delim)})"
                        result["sheets"] = {"Data": df}
                    else:
                        raise ValueError
                except Exception:
                    result["source_type"] = "Text/Log"
                    result["sheets"] = parse_text_log(text)

        else:
            # Unknown: try CSV then text
            text = raw.decode(enc, errors="replace")
            try:
                result["sheets"] = parse_csv(raw)
                result["source_type"] = "Auto-detected CSV"
            except Exception:
                result["source_type"] = "Unknown text"
                result["sheets"] = parse_text_log(text)

    except Exception as exc:
        result["errors"].append(f"Conversion error: {exc}")
        result["errors"].append(traceback.format_exc())

    return _post_process(result)


# ── Post-processing: clean + type-cast ───────────────────────────────────────

def _post_process(result: dict[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, pd.DataFrame] = {}
    for sheet_name, df in result["sheets"].items():
        if not isinstance(df, pd.DataFrame) or df.empty:
            cleaned[sheet_name] = df
            continue
        df = _clean_df(df)
        cleaned[sheet_name] = df
    result["sheets"] = cleaned
    return result


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    # Normalise column names
    df.columns = [
        re.sub(r"\s+", "_", str(c).strip()).lower() if str(c).strip() else f"col_{i}"
        for i, c in enumerate(df.columns)
    ]
    # Drop completely empty rows/cols
    df = df.dropna(how="all").reset_index(drop=True)
    df = df.loc[:, ~(df.isna().all())]
    # Deduplicate
    df = df.drop_duplicates().reset_index(drop=True)
    # Smart type casting
    for col in df.columns:
        df[col] = _smart_cast(df[col])
    return df


def _smart_cast(series: pd.Series) -> pd.Series:
    # Try numeric
    if series.dtype == object:
        try:
            numeric = pd.to_numeric(series.str.replace(",", "", regex=False), errors="coerce")
            if numeric.notna().sum() / max(series.notna().sum(), 1) > 0.8:
                return numeric
        except AttributeError:
            pass
        # Try datetime
        try:
            dt = pd.to_datetime(series, infer_datetime_format=True, errors="coerce")
            if dt.notna().sum() / max(series.notna().sum(), 1) > 0.7:
                return dt
        except Exception:
            pass
        # Try boolean
        bool_map = {"true": True, "false": False, "yes": True, "no": False, "1": True, "0": False}
        lower = series.str.lower().str.strip()
        if lower.dropna().isin(bool_map).all():
            return lower.map(bool_map)
    return series


# ── Excel writer ──────────────────────────────────────────────────────────────

def to_excel(sheets: dict[str, pd.DataFrame], include_summary: bool = True) -> bytes:
    """Write sheets to an Excel workbook with formatting."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Summary sheet
        if include_summary:
            summary_rows = []
            for name, df in sheets.items():
                for col in df.columns:
                    summary_rows.append({
                        "Sheet": name,
                        "Column": col,
                        "dtype": str(df[col].dtype),
                        "Non-Null": df[col].notna().sum(),
                        "Null": df[col].isna().sum(),
                        "Unique": df[col].nunique(),
                        "Sample": str(df[col].dropna().iloc[0]) if df[col].notna().any() else "",
                    })
            if summary_rows:
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name="📋 Summary", index=False)

        # Data sheets
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    # Apply formatting with openpyxl
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    _format_workbook(wb)
    out2 = io.BytesIO()
    wb.save(out2)
    return out2.getvalue()


def _format_workbook(wb):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1A1F2E")
    HEADER_FONT = Font(bold=True, color="00E5FF", name="Arial", size=10)
    CELL_FONT = Font(name="Arial", size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT = Alignment(horizontal="left", vertical="center")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = CELL_FONT
                cell.alignment = LEFT

        # Auto-width
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"

    return wb
